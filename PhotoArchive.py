from genericpath import isdir
import pandas as pd
from openpyxl import load_workbook
import os
import shutil
import os.path
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import filedialog

# Be sure to understand readme file before starting your tranfer.
# Some of the sections must be edited before you start.
# Your file system according to your items list must be prepared before start this process.
# If your naming logic is different than what I wrote, then you need to change the
# searching pattern. Don't hesitate to contact me for any question!

root = tk.Tk()
root.withdraw()

# Choose the photo pool you want to move.
directory_path = filedialog.askdirectory(
    title="Choose the photo pool you want to move")
print(directory_path)

# Choose where the photos will be moved to.
new_path = filedialog.askdirectory(
    title="#Choose where the photos will be moved to.")
print(new_path)

# Chose the master excel file.
file_path = filedialog.askopenfilename(title="Chose the master excel file")
print(file_path)

# Choose the report paper.
archive_excel = filedialog.askopenfilename(title="Choose the report paper.")
print(archive_excel)

archive_count = 2
archive = openpyxl.load_workbook(archive_excel)
archive_sheet = archive['main']

excel_file = pd.ExcelFile(file_path)
sheets = excel_file.sheet_names
print(sheets)

dirs_list = []


def listdirs(rootdir):
    for file in os.listdir(rootdir):
        d = rootdir + "/"+file
        if os.path.isdir(d):
            print(d)
            dirs_list.append(d)

            listdirs(d)


shm_number = ""

listdirs(directory_path)

archive_count += 1
for dir in dirs_list:
    print("WORKING ON THE DIRECTORY : " + dir + " ...")
    for filename in os.listdir(dir):
        shm_number = ''
        print("-----------------------------------------------------------------------------------------------")
        print("WORKING ON THE ITEM : " + filename + " ...")

        first_line = 0
        find = False

        # You might be wondering why this part exists, in Turkish we have 'i' and 'ı',
        # This is why we needed the code below.
        for i in range(0, len(filename)):
            if filename[i] == "_" or filename[i] == "." or filename[i] == "-" or filename[i] == " ":
                cat = filename[0:i]
                if cat.startswith("i"):
                    cat = "İ"+cat[1:]
                elif cat.startswith("ı"):
                    cat = "İ"+cat[1:]
                elif cat.startswith("I"):
                    cat = "İ"+cat[1:]
                cat = cat.upper()

                # In here file name is iterated and if any of the following ('_','.','-',' ') is spotted,
                # Then the catalog number for it is decided.
                # shm is the abbraviation for the museum that I was working at when I wrote this program.
                for j in range(i+1, len(filename)):
                    if filename[j] == "_" or filename[j] == "." or filename[j] == "-" or filename[j] == " ":
                        shm_number = cat + "_" + filename[i+1: j]
                        find = True

                        print(shm_number)
                        break
            if find == True:
                break

        for sheet in sheets:
            print("WORKING ON THE SHEET : " + sheet + " ...")
            page = excel_file.parse(sheet)
            lenght, widht = page.shape

            # iterating through the sheet for items
            for i in range(0, lenght):
                photo_check = page['Sira3'][i]

                # checking if there is a match
                if shm_number == str(photo_check):
                    print("BINGO!! I find a match.. now I am moving it!!")

                    try:
                        # DECIDING FOR FILE TYPE
                        # Program repeats itself for different types of files.
                        # If you want to use it for just one file type then bend it to your will.
                        movedTo = new_path+"/" + sheet + "/" + page['Sira1'][i]
                        # TIF file
                        if filename[-3:] == "tif":
                            extention = ".tif"
                            l = 1
                            # The path written here should be changed according to your file system.
                            # In this version of the program the photos are diveden into there seperate format
                            # .raw ones goes to untouched folder, tif and .jpg goes to touched folder.
                            if os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/İşlenmiş/TIF/" + page['Sira1'][i] + extention):
                                while os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/İşlenmiş/TIF/" + page['Sira1'][i]+"_" + str(l) + extention):
                                    l = l+1
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/İşlenmiş/TIF/" + page['Sira1'][i] + "_" + str(l) + extention)
                                # Below is what will be written to your report.
                                # Change it according to your desire if want it to write something else
                                # In this version it is old name of the file, where it is moved to and what it the new name of file
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' + str(archive_count)
                                              ] = movedTo + "/İşlenmiş/TIF/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i] + "_" + str(l)+extention
                                archive_count += 1
                            else:
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/İşlenmiş/TIF/" + page['Sira1'][i] + extention)
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' + str(archive_count)
                                              ] = movedTo + "/İşlenmiş/TIF/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i]+extention
                                archive_count += 1

                        # JPG file
                        if filename[-3:] == "jpg" or filename[-3:] == "JPG" or filename[-3:] == "jpeg" or filename[-3:] == "JPEG":
                            extention = ".jpg"
                            l = 1
                            if os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/İşlenmiş/JPG/" + page['Sira1'][i] + extention):
                                while os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/İşlenmiş/JPG/" + page['Sira1'][i]+"_" + str(l) + extention):
                                    l = l+1
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/İşlenmiş/JPG/" + page['Sira1'][i] + "_" + str(l) + extention)
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' + str(archive_count)
                                              ] = movedTo + "/İşlenmiş/JPG/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i] + "_" + str(l)+extention
                                archive_count += 1
                            else:
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/İşlenmiş/JPG/" + page['Sira1'][i] + extention)
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' + str(archive_count)
                                              ] = movedTo + "/İşlenmiş/JPG/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i]+extention
                                archive_count += 1

                        # PNG file
                        if filename[-3:] == "png":
                            extention = ".png"
                            l = 1
                            if os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/İşlenmiş/JPG/" + page['Sira1'][i] + extention):
                                while os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/İşlenmiş/JPG/" + page['Sira1'][i]+"_" + str(l) + extention):
                                    l = l+1
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/İşlenmiş/JPG/" + page['Sira1'][i] + "_" + str(l) + extention)
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' + str(archive_count)
                                              ] = movedTo + "/İşlenmiş/JPG/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i] + "_" + str(l)+extention
                                archive_count += 1
                            else:
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/İşlenmiş/JPG/" + page['Sira1'][i] + extention)
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' + str(archive_count)
                                              ] = movedTo + "/İşlenmiş/JPG/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i]+extention
                                archive_count += 1

                        # RAW file
                        if filename[-3:] == "raw" or filename[-3:] == "RAW":
                            extention = ".raw"
                            l = 1
                            if os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/Ham/" + page['Sira1'][i] + extention):
                                while os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/Ham/" + page['Sira1'][i] + "_" + str(l) + extention):
                                    l = l+1
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/Ham/" + page['Sira1'][i] + "_" + str(l) + extention)
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' +
                                              str(archive_count)] = movedTo + "/Ham/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i] + "_" + str(l)+extention
                                archive_count += 1
                            else:
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/Ham/" + page['Sira1'][i] + extention)
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' +
                                              str(archive_count)] = movedTo + "/Ham/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i]+extention
                                archive_count += 1

                        # NEF file
                        if filename[-3:] == "nef" or filename[-3:] == "NEF":
                            extention = ".nef"
                            l = 1
                            if os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/Ham/" + page['Sira1'][i] + extention):
                                while os.path.exists(new_path+"/" + sheet + "/" + page['Sira1'][i] + "/Ham/" + page['Sira1'][i] + "_" + str(l) + extention):
                                    l = l+1
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/Ham/" + page['Sira1'][i] + "_" + str(l) + extention)
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' +
                                              str(archive_count)] = movedTo + "/Ham/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i] + "_" + str(l)+extention
                                archive_count += 1
                            else:
                                shutil.move(dir+"/"+filename, new_path+"/" + sheet + "/" +
                                            page['Sira1'][i] + "/Ham/" + page['Sira1'][i] + extention)
                                archive_sheet['A' +
                                              str(archive_count)] = filename
                                archive_sheet['B' +
                                              str(archive_count)] = movedTo + "/Ham/"
                                archive_sheet['C' + str(archive_count)
                                              ] = page['Sira1'][i]+extention
                                archive_count += 1

                    except:
                        print("Can't move this one, There is a mistake!")
                        archive_sheet['A' + str(archive_count)] = filename
                        archive_sheet['B' + str(archive_count)] = "Can't move"
                        archive_sheet['C' + str(archive_count)] = "Error!!!"
                    else:
                        continue

archive.save(archive_excel)
